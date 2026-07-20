from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

if __package__ in (None, ""):
    process_dir = Path(__file__).resolve().parents[1]
    if str(process_dir) not in sys.path:
        sys.path.insert(0, str(process_dir))

GENERATOR_VERSION = "0.3.0-f-openpyxl-portable"


def main() -> None:
    manifest_path, backup_label = _parse_args(sys.argv[1:])
    run_export(manifest_path, backup_label)


def run_export(
    manifest_path: Path | str,
    backup_label: str | None = None,
    thread_delivery_dir: Path | str | None = None,
) -> dict[str, Any]:
    """Export exactly the D3 manifest supplied by J or the caller.

    The project package is read only from the supplied manifest.  No historical
    package, source directory, Node binary, or user-specific runtime is used.
    """
    from f_exporter.export_model_d3 import append_f_outputs_to_file_inventory, build_d3_export_model
    from f_exporter.workbook import D3_SHEET_ORDER, write_workbook

    manifest_path = _require_file(Path(manifest_path), "D3 manifest")
    manifest = _read_json(manifest_path)
    if not isinstance(manifest, dict):
        raise ValueError(f"D3 manifest must be a JSON object: {manifest_path}")

    outputs = manifest.get("outputs")
    if not isinstance(outputs, dict):
        raise ValueError(f"D3 manifest outputs must be an object: {manifest_path}")
    cards = _read_json(_required_output_file(outputs, "parameter_cards", manifest_path))
    source_refs = _read_json(_required_output_file(outputs, "source_refs", manifest_path))
    issues = _read_json(_required_output_file(outputs, "issues", manifest_path))
    if not all(isinstance(rows, list) for rows in (cards, source_refs, issues)):
        raise ValueError(f"D3 export JSON inputs must each be arrays: {manifest_path}")

    project_package_value = manifest.get("project_package")
    if not isinstance(project_package_value, str) or not project_package_value.strip():
        raise ValueError(f"D3 manifest is missing project_package: {manifest_path}")
    project_package = Path(project_package_value).expanduser()
    if not project_package.is_dir():
        raise FileNotFoundError(f"D3 project package does not exist: {project_package}")

    output_stem = _output_stem_from_project_package(project_package)
    raw_files = _list_files(project_package / "01_原始询价文件")
    translation_files = _list_files(project_package / "02_中文翻译文件")

    f_system_dir = project_package / "系统数据" / "参数汇总结果_F"
    xlsx_dir = project_package / "03_参数汇总表"
    report_dir = project_package / "04_处理报告"
    for directory in (f_system_dir, xlsx_dir, report_dir):
        directory.mkdir(parents=True, exist_ok=True)

    backup_dir = None
    backup_copied: list[Path] = []
    if backup_label:
        backup_dir, backup_copied = _backup_existing_outputs(project_package, output_stem, backup_label)

    xlsx_path = xlsx_dir / f"参数汇总表_{output_stem}.xlsx"
    source_report_path = report_dir / f"来源定位报告_{output_stem}.txt"
    review_report_path = report_dir / f"待复核问题报告_{output_stem}.txt"
    export_tables_path = f_system_dir / "export_tables_f.json"
    processing_report_path = f_system_dir / "f_processing_report.txt"
    f_manifest_path = f_system_dir / "f_thread_manifest.json"
    output_paths = {
        "xlsx": str(xlsx_path),
        "source_report": str(source_report_path),
        "review_report": str(review_report_path),
        "manifest": str(f_manifest_path),
        "export_tables": str(export_tables_path),
        "processing_report": str(processing_report_path),
    }

    model = build_d3_export_model(manifest, cards, source_refs, issues, raw_files, translation_files)
    source_report_path.write_text(_source_report(model, output_stem), encoding="utf-8")
    review_report_path.write_text(_review_report(model, output_stem), encoding="utf-8")
    export_tables_path.write_text(json.dumps(model, ensure_ascii=False, indent=2), encoding="utf-8")
    processing_report_path.write_text(_processing_report(manifest_path, manifest, model, output_paths, backup_dir), encoding="utf-8")
    f_manifest_path.write_text("{}", encoding="utf-8")

    write_workbook(model, xlsx_path, D3_SHEET_ORDER)
    append_f_outputs_to_file_inventory(model, output_paths)
    export_tables_path.write_text(json.dumps(model, ensure_ascii=False, indent=2), encoding="utf-8")
    processing_report_path.write_text(_processing_report(manifest_path, manifest, model, output_paths, backup_dir), encoding="utf-8")
    write_workbook(model, xlsx_path, D3_SHEET_ORDER)

    copied = _copy_to_explicit_delivery(
        thread_delivery_dir,
        [xlsx_path, source_report_path, review_report_path, export_tables_path, processing_report_path],
    )
    validation = _validate_outputs(output_paths, list(D3_SHEET_ORDER))
    f_manifest = {
        "thread": "F",
        "generator_version": GENERATOR_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "project_package": str(project_package),
        "input_manifest": str(manifest_path),
        "input_d3_version": manifest.get("generator_version"),
        "input_d3_statistics": manifest.get("statistics", {}),
        "outputs": output_paths,
        "copied_to_thread_delivery": copied,
        "backup": {
            "backup_dir": str(backup_dir) if backup_dir else "",
            "copied_files": [str(path) for path in backup_copied],
        },
        "statistics": {
            "card_count": len(model["参数卡片总览"]),
            "parameter_row_count": len(model["参数明细"]),
            "source_row_count": len(model["来源定位"]),
            "issue_row_count": len(model["待复核问题"]),
            "download_artifact_count": len(model["文件清单"]),
        },
        "validation": validation,
    }
    f_manifest_path.write_text(json.dumps(f_manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    if thread_delivery_dir:
        manifest_copy = Path(thread_delivery_dir) / f_manifest_path.name
        manifest_copy.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(f_manifest_path, manifest_copy)
        f_manifest["copied_to_thread_delivery"].append(str(manifest_copy))
        f_manifest_path.write_text(json.dumps(f_manifest, ensure_ascii=False, indent=2), encoding="utf-8")
        shutil.copy2(f_manifest_path, manifest_copy)

    print(json.dumps(f_manifest["statistics"], ensure_ascii=False))
    return f_manifest


def _parse_args(args: list[str]) -> tuple[Path, str | None]:
    parser = argparse.ArgumentParser(
        prog=Path(__file__).name,
        description="Generate F summary and review reports from an explicit D3 manifest.",
    )
    parser.add_argument(
        "d3_manifest_path",
        help="Path to the required d3_thread_manifest.json. No historical sample fallback is used.",
    )
    parser.add_argument(
        "--backup-label",
        help="Optional label for backing up existing F outputs in the target project package.",
    )
    parsed = parser.parse_args(args)
    return Path(parsed.d3_manifest_path), parsed.backup_label


def _output_stem_from_project_package(project_package: Path) -> str:
    name = project_package.name
    return name[3:] if name.startswith("项目_") else name


def _backup_existing_outputs(project_package: Path, output_stem: str, backup_label: str) -> tuple[Path, list[Path]]:
    f_system_dir = project_package / "系统数据" / "参数汇总结果_F"
    xlsx_dir = project_package / "03_参数汇总表"
    report_dir = project_package / "04_处理报告"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = f_system_dir / "历史版本" / f"{backup_label}_{timestamp}"
    backup_dir.mkdir(parents=True, exist_ok=True)
    candidates = [
        f_system_dir / "f_thread_manifest.json",
        f_system_dir / "export_tables_f.json",
        f_system_dir / "f_processing_report.txt",
        xlsx_dir / f"参数汇总表_{output_stem}.xlsx",
        report_dir / f"来源定位报告_{output_stem}.txt",
        report_dir / f"待复核问题报告_{output_stem}.txt",
    ]
    copied = []
    for path in candidates:
        if path.exists():
            destination = backup_dir / path.name
            shutil.copy2(path, destination)
            copied.append(destination)
    return backup_dir, copied


def _required_output_file(outputs: dict[str, Any], key: str, manifest_path: Path) -> Path:
    value = outputs.get(key)
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"D3 manifest is missing outputs.{key}: {manifest_path}")
    return _require_file(Path(value), f"D3 outputs.{key}")


def _require_file(path: Path, label: str) -> Path:
    if not path.is_file():
        raise FileNotFoundError(f"{label} does not exist or is not a file: {path}")
    return path


def _read_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"JSON cannot be parsed: {path}") from exc


def _list_files(directory: Path) -> list[dict[str, Any]]:
    if not directory.exists():
        return []
    rows = []
    for path in sorted(item for item in directory.iterdir() if item.is_file()):
        rows.append({"name": path.name, "path": str(path), "exists": path.exists(), "size_kb": round(path.stat().st_size / 1024, 1)})
    return rows


def _copy_to_explicit_delivery(delivery_dir: Path | str | None, paths: list[Path]) -> list[str]:
    if delivery_dir is None:
        return []
    destination_dir = Path(delivery_dir)
    destination_dir.mkdir(parents=True, exist_ok=True)
    copied = []
    for path in paths:
        destination = destination_dir / path.name
        shutil.copy2(path, destination)
        copied.append(str(destination))
    return copied


def _source_report(model: dict[str, list[dict[str, Any]]], output_stem: str) -> str:
    grouped: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for row in model["来源定位"]:
        grouped.setdefault(row["Tag No."], {}).setdefault(row["字段 key"], []).append(row)
    lines = [f"来源定位报告_{output_stem}", ""]
    lines.append(f"来源行数：{len(model['来源定位'])}")
    lines.append("说明：本报告按 Tag No. 和字段列出 D3 supporting_sources，F 不重新判断参数含义。")
    for tag in sorted(grouped):
        lines.extend(["", f"Tag No.: {tag}"])
        for field_key in sorted(grouped[tag]):
            lines.append(f"  字段：{field_key}")
            for source in grouped[tag][field_key]:
                lines.append(
                    "    - "
                    f"{source['source_ref_id']} | {source['来源短标识']} | {source['文件名']} | "
                    f"page {source['PDF 页码']} | {source['表格/行列信息']} | "
                    f"method={source['抽取方式']} | confidence={source['置信度']} | verified={source['evidence_verified']}"
                )
                if source["原文片段"]:
                    lines.append(f"      原文：{source['原文片段']}")
                if source["中文译文片段"]:
                    lines.append(f"      译文：{source['中文译文片段']}")
    return "\n".join(lines) + "\n"


def _review_report(model: dict[str, list[dict[str, Any]]], output_stem: str) -> str:
    rows = model["待复核问题"]
    by_code: dict[str, int] = {}
    for row in rows:
        by_code[row["问题代码"]] = by_code.get(row["问题代码"], 0) + 1
    lines = [f"待复核问题报告_{output_stem}", ""]
    lines.append(f"待复核问题总数：{len(rows)}")
    lines.append("问题代码统计：")
    for code, count in sorted(by_code.items()):
        lines.append(f"- {code}：{count}")
    lines.extend(["", "问题明细："])
    for index, row in enumerate(rows, start=1):
        lines.append(
            f"{index}. [{row['风险等级']}] {row['问题代码']} | Tag={row['Tag No.'] or '-'} | "
            f"字段={row['字段名'] or '-'} | 文件={row['文件名'] or '-'} | {row['问题说明']} | 建议：{row['建议复核动作']}"
        )
    return "\n".join(lines) + "\n"


def _processing_report(
    manifest_path: Path,
    manifest: dict[str, Any],
    model: dict[str, list[dict[str, Any]]],
    output_paths: dict[str, str],
    backup_dir: Path | None,
) -> str:
    lines = [
        "F D3 数据汇总处理报告",
        "",
        f"输入 D3 manifest：{manifest_path}",
        f"D3 版本：{manifest.get('generator_version')}",
        f"D3 统计：{json.dumps(manifest.get('statistics', {}), ensure_ascii=False)}",
        "",
        "F 输出统计：",
    ]
    for name, rows in model.items():
        lines.append(f"- {name}：{len(rows)} 行")
    lines.extend(["", "输出文件："])
    for key, path in output_paths.items():
        lines.append(f"- {key}: {path}")
    if backup_dir:
        lines.extend(["", f"旧 F 输出备份目录：{backup_dir}"])
    lines.extend(["", "边界说明：本次未读取 D2/C2 旧样例、未生成 selection_draft_input_package.json、未修改 D3/B/J/A/Reference。"])
    return "\n".join(lines) + "\n"


def _validate_outputs(output_paths: dict[str, str], expected_sheets: list[str]) -> dict[str, Any]:
    paths_exist = {
        key: Path(path).exists() and Path(path).stat().st_size > 0
        for key, path in output_paths.items()
        if key != "manifest"
    }
    try:
        sheet_names, formula_error_cells = _inspect_workbook(Path(output_paths["xlsx"]))
        return {
            "json_readable": True,
            "excel_openable": True,
            "key_sheets_present": all(sheet in sheet_names for sheet in expected_sheets),
            "sheet_names": sheet_names,
            "formula_error_cell_count": formula_error_cells,
            "output_files_exist": paths_exist,
        }
    except Exception as exc:
        return {
            "json_readable": True,
            "excel_openable": False,
            "key_sheets_present": False,
            "sheet_names": [],
            "formula_error_cell_count": None,
            "output_files_exist": paths_exist,
            "excel_error": str(exc),
        }


def _inspect_workbook(xlsx_path: Path) -> tuple[list[str], int]:
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
    try:
        error_cells = sum(
            1
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.data_type == "e"
        )
        return workbook.sheetnames, error_cells
    finally:
        workbook.close()


if __name__ == "__main__":
    main()
