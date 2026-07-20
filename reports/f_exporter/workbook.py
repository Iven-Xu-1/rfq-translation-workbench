from __future__ import annotations

import json
import os
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


D3_SHEET_ORDER = (
    "项目总览",
    "参数卡片总览",
    "参数明细",
    "来源定位",
    "待复核问题",
    "文件清单",
)

LEGACY_SHEET_ORDER = (
    "参数卡片总览",
    "参数明细",
    "来源索引",
    "待复核问题",
    "模板字段对照",
)

_MANUAL_COLUMNS = {
    "人工复核结论",
    "人工修正值",
    "人工备注",
    "人工确认来源",
    "人工处理结论",
}
_MANUAL_LIST_COLUMNS = {"人工复核结论", "人工确认来源", "人工处理结论"}
_MANUAL_OPTIONS = "确认正确,修改参数,选择来源,标记不适用,保留疑问"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_EDITABLE_FILL = PatternFill("solid", fgColor="FFF7E6")
_REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")
_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Arial", size=10, color="000000")
_REVIEW_FONT = Font(name="Arial", size=10, bold=True, color="9C0006")
_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def write_workbook(
    tables: Mapping[str, Sequence[Mapping[str, Any]]],
    xlsx_path: Path,
    sheet_order: Sequence[str] = D3_SHEET_ORDER,
) -> None:
    """Write the review workbook without relying on Node or a development runtime."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in sheet_order:
        _write_sheet(workbook, sheet_name, tables.get(sheet_name, []))

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = xlsx_path.with_name(f".{xlsx_path.stem}.tmp.xlsx")
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, xlsx_path)
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()


def _write_sheet(workbook: Workbook, sheet_name: str, rows: Sequence[Mapping[str, Any]]) -> None:
    if not isinstance(rows, Sequence):
        raise ValueError(f"Workbook table {sheet_name} must be a list of rows.")
    normalized_rows = []
    for row in rows:
        if not isinstance(row, Mapping):
            raise ValueError(f"Workbook table {sheet_name} contains a non-object row.")
        normalized_rows.append(row)

    headers = list(normalized_rows[0].keys()) if normalized_rows else ["无数据"]
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=str(header))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = _HEADER_ALIGNMENT
    worksheet.row_dimensions[1].height = 30

    for row_index, row in enumerate(normalized_rows, start=2):
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=_cell_value(row.get(header)))
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = _BODY_ALIGNMENT

    _style_manual_review_columns(worksheet, headers)
    _style_review_flag_column(worksheet, headers)
    _set_column_widths(worksheet, headers)


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _style_manual_review_columns(worksheet, headers: list[str]) -> None:
    if worksheet.max_row < 2:
        return
    for column_index, header in enumerate(headers, start=1):
        if header not in _MANUAL_COLUMNS:
            continue
        column_letter = get_column_letter(column_index)
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).fill = _EDITABLE_FILL
        if header in _MANUAL_LIST_COLUMNS:
            validation = DataValidation(type="list", formula1=f'"{_MANUAL_OPTIONS}"', allow_blank=True)
            worksheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}{worksheet.max_row}")


def _style_review_flag_column(worksheet, headers: list[str]) -> None:
    if worksheet.max_row < 2 or "是否待复核" not in headers:
        return
    column_letter = get_column_letter(headers.index("是否待复核") + 1)
    worksheet.conditional_formatting.add(
        f"{column_letter}2:{column_letter}{worksheet.max_row}",
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("是",{column_letter}2))'],
            fill=_REVIEW_FILL,
            font=_REVIEW_FONT,
        ),
    )


def _set_column_widths(worksheet, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        longest = len(str(header))
        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            longest = max(longest, len(str(value)) if value is not None else 0)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(longest + 2, 11), 48)
